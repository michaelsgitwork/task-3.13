import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Get webpage
url = "https://test-scrape-site.onrender.com/planets.html"
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# Get text 
text = soup.get_text()

# Split 
planet_parts = text.split('\n')

Excel file
wb = Workbook()
ws = wb.active
ws.title = "Planets"

# headers
ws['A1'] = 'Name'
ws['B1'] = 'Position'
ws['C1'] = 'Diameter'
ws['D1'] = 'Moons'
ws['E1'] = 'Atmosphere'
ws['F1'] = 'Fun Fact'

#  add data
planets = []
current_planet = {}

for line in planet_parts:
    line = line.strip()
    
    if line.startswith(('Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune', 'Pluto', 'Eris')):
        if current_planet:
            planets.append(current_planet)
        current_planet = {'Name': line}
    
    elif 'Position:' in line:
        current_planet['Position'] = line.replace('Position:', '').strip()
    elif 'Diameter:' in line:
        current_planet['Diameter'] = line.replace('Diameter:', '').strip()
    elif 'Moons:' in line:
        current_planet['Moons'] = line.replace('Moons:', '').strip()
    elif 'Atmosphere:' in line:
        current_planet['Atmosphere'] = line.replace('Atmosphere:', '').strip()
    elif 'Fun Fact:' in line:
        current_planet['Fun Fact'] = line.replace('Fun Fact:', '').strip()

# Add last planet
if current_planet:
    planets.append(current_planet)

# Write to Excel
for i, planet in enumerate(planets, 2):
    ws[f'A{i}'] = planet.get('Name', '')
    ws[f'B{i}'] = planet.get('Position', '')
    ws[f'C{i}'] = planet.get('Diameter', '')
    ws[f'D{i}'] = planet.get('Moons', '')
    ws[f'E{i}'] = planet.get('Atmosphere', '')
    ws[f'F{i}'] = planet.get('Fun Fact', '')

# Save file
wb.save('planets_data.xlsx')

print(f"Found {len(planets)} planets!")
print("Data saved to planets_data.xlsx")
